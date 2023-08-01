from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter

from .style import DefaultStyle


class RowWriter:
    def __init__(self):
        self.wb = Workbook(write_only=True)
        self.__row_list = []

    def create_sheet(self, title: str, protection: bool = False):
        self.ws = self.wb.create_sheet(title)
        self.ws.protection.sheet = protection

    def switch_current_sheet(self, sheet: str):
        self.ws = self.wb[sheet]

    def print_row_list(self, mode='value'):
        if mode == 'value':
            print('row_list = [', end='')
            for item in self.__row_list:
                if item == self.__row_list[-1]:
                    print(f'{item.value}', end='')
                else:
                    print(f'{item.value}, ', end='')
            print(']')
        else:
            print(self.__row_list)

    def get_current_sheet(self):
        return self.ws

    def row_append(self, value, style=DefaultStyle(), **kwargs):
        cell = WriteOnlyCell(self.ws, value=value)
        cell.font = style.font
        cell.fill = style.fill
        cell.alignment = style.ali
        cell.border = style.border
        cell.protection = style.protection

        self.__row_list.append(cell)

    def row_append_list(self, data: list, style=DefaultStyle()):
        for val in data:
            cell = WriteOnlyCell(self.ws, value=val)
            cell.font = style.font
            cell.fill = style.fill
            cell.alignment = style.ali
            cell.border = style.border
            cell.protection = style.protection
            cell.number_format = style.number_format

            self.__row_list.append(cell)

    def create_row(self):
        self.ws.append(self.__row_list)
        self.__row_list = []

    def set_cell_width(self, col: int, width: int):
        self.ws.column_dimensions[get_column_letter(col)].width = width

    def set_cell_height(self, row: int, height: int):
        self.ws.row_dimensions[row].height = height

    def save(self, name='output.xlsx'):
        self.wb.save(name)
