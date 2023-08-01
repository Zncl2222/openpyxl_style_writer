from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side


class DefaultStyle:
    # params
    font_params = None
    fill_params = None
    ali_params = None
    border_params = None

    # font
    font_size = 14
    font_name = 'Calibri'
    font_bold = False
    font_italic = False
    font_underline = 'none'
    font_strike = False
    font_vertAlign = None
    font_color = '000000'

    # fill
    fill_pattern = 'solid'
    fill_color = 'fcfcfc'

    # alignment
    ali_horizontal = 'center'
    ali_vertical = 'center'
    ali_text_rotation = 0
    ali_wrap_text = False
    ali_shrink_to_fit = False
    ali_indent = 0

    # border
    border_style_top = None
    border_style_right = None
    border_style_left = None
    border_style_bottom = None
    border_color_top = 'ff000000'
    border_color_right = 'ff000000'
    border_color_left = 'ff000000'
    border_color_bottom = 'ff000000'

    # protect
    protect = False
    protection = Protection(locked=False)

    # format
    number_format = 'General'

    font = Font(
        size=font_size,
        name=font_name,
        bold=font_bold,
        color=font_color,
    )

    fill = PatternFill(fill_pattern, fgColor=fill_color)

    ali = Alignment(
        horizontal=ali_horizontal,
        vertical=ali_vertical,
        wrap_text=ali_wrap_text,
    )

    border = Border(
        top=Side(
            border_style=border_style_top,
            color=border_color_top,
        ),
        right=Side(
            border_style=border_style_right,
            color=border_color_right,
        ),
        left=Side(
            border_style=border_style_left,
            color=border_color_left,
        ),
        bottom=Side(
            border_style=border_style_bottom,
            color=border_color_bottom,
        ),
    )

    @classmethod
    def set_default(cls, **kwargs):
        # font settings
        if kwargs.get('font_params'):
            cls.font_params = kwargs.get('font_params')
        else:
            cls.font_name = kwargs.get('font_name', cls.font_name)
            cls.font_color = kwargs.get('font_color', cls.font_color)
            cls.font_size = kwargs.get('font_size', cls.font_size)
            cls.font_bold = kwargs.get('font_bold', cls.font_bold)

        # fill settings
        if kwargs.get('fill_params'):
            cls.fill_params = kwargs.get('fill_params')
        else:
            cls.fill_color = kwargs.get('fill_color', cls.fill_color)

        # alignment settings
        if kwargs.get('ali_params'):
            cls.ali_params = kwargs.get('ali_params')
        else:
            cls.ali_horizontal = kwargs.get('ali_horizontal', cls.ali_horizontal)
            cls.ali_vertical = kwargs.get('ali_vertical', cls.ali_vertical)
            cls.ali_wrap_text = kwargs.get('ali_wrap_text', cls.ali_wrap_text)

        # border settings
        if kwargs.get('border_params'):
            cls.border_params = kwargs.get('border_params')
        else:
            cls.border_style_top = kwargs.get('border_style_top', cls.border_style_top)
            cls.border_style_right = kwargs.get('border_style_right', cls.border_style_right)
            cls.border_style_left = kwargs.get('border_style_left', cls.border_style_left)
            cls.border_style_bottom = kwargs.get('border_style_bottom', cls.border_style_bottom)
            cls.border_color_top = kwargs.get('border_color_top', cls.border_color_top)
            cls.border_color_right = kwargs.get('border_color_right', cls.border_color_right)
            cls.border_color_left = kwargs.get('border_color_left', cls.border_color_left)
            cls.border_color_bottom = kwargs.get('border_color_bottom', cls.border_color_bottom)

        # protection setting
        if kwargs.get('protect'):
            cls.protect = kwargs.get('protect')

        # number format setting
        if kwargs.get('number_format'):
            cls.number_format = kwargs.get('number_format')

        cls.apply_settings()

    @classmethod
    def apply_settings(cls):
        if cls.font_params:
            cls.font = Font(**cls.font_params)
        else:
            cls.font = Font(
                size=cls.font_size,
                name=cls.font_name,
                bold=cls.font_bold,
                color=cls.font_color,
            )

        if cls.fill_params:
            cls.fill = PatternFill(**cls.fill_params)
        else:
            cls.fill = PatternFill(cls.fill_pattern, fgColor=cls.fill_color)

        if cls.ali_params:
            cls.ali = Alignment(**cls.ali_params)
        else:
            cls.ali = Alignment(
                horizontal=cls.ali_horizontal,
                vertical=cls.ali_vertical,
                wrap_text=cls.ali_wrap_text,
            )

        if cls.border_params:
            cls.border = Border(**cls.border_params)
        else:
            cls.border = Border(
                top=Side(
                    border_style=cls.border_style_top,
                    color=cls.border_color_top,
                ),
                right=Side(
                    border_style=cls.border_style_right,
                    color=cls.border_color_right,
                ),
                left=Side(
                    border_style=cls.border_style_left,
                    color=cls.border_color_left,
                ),
                bottom=Side(
                    border_style=cls.border_style_bottom,
                    color=cls.border_color_bottom,
                ),
            )

        if cls.protect is True:
            cls.protection = Protection(locked=True)


class CustomStyle(DefaultStyle):
    font_params = None
    fill_params = None
    ali_params = None
    border_params = None

    def __init__(self, **kwargs):
        # font settings
        if kwargs.get('font_params'):
            self.font_params = kwargs.get('font_params')
        else:
            self.font_name = kwargs.get('font_name', self.font_name)
            self.font_color = kwargs.get('font_color', self.font_color)
            self.font_size = kwargs.get('font_size', self.font_size)
            self.font_bold = kwargs.get('font_bold', self.font_bold)

        # fill settings
        if kwargs.get('fill_params'):
            self.fill_params = kwargs.get('fill_params')
        else:
            self.fill_color = kwargs.get('fill_color', self.fill_color)

        # alignment settings
        if kwargs.get('ali_params'):
            self.ali_params = kwargs.get('ali_params')
        else:
            self.ali_horizontal = kwargs.get('ali_horizontal', self.ali_horizontal)
            self.ali_vertical = kwargs.get('ali_vertical', self.ali_vertical)
            self.ali_wrap_text = kwargs.get('ali_wrap_text', self.ali_wrap_text)

        # border settings
        if kwargs.get('border_params'):
            self.border_params = kwargs.get('border_params')
        else:
            self.border_style_top = kwargs.get('border_style_top', self.border_style_top)
            self.border_style_right = kwargs.get('border_style_right', self.border_style_right)
            self.border_style_left = kwargs.get('border_style_left', self.border_style_left)
            self.border_style_bottom = kwargs.get('border_style_bottom', self.border_style_bottom)
            self.border_color_top = kwargs.get('border_color_top', self.border_color_top)
            self.border_color_right = kwargs.get('border_color_right', self.border_color_right)
            self.border_color_left = kwargs.get('border_color_left', self.border_color_left)
            self.border_color_bottom = kwargs.get('border_color_bottom', self.border_color_bottom)

        if kwargs.get('protect'):
            self.protect = True

        if kwargs.get('number_format'):
            self.number_format = kwargs.get('number_format')

        self.apply_settings()

    def apply_settings(self):
        if self.font_params:
            self.font = Font(**self.font_params)
        else:
            self.font = Font(
                size=self.font_size,
                name=self.font_name,
                bold=self.font_bold,
                color=self.font_color,
            )

        if self.fill_params:
            self.fill = PatternFill(**self.fill_params)
        else:
            self.fill = PatternFill(self.fill_pattern, fgColor=self.fill_color)

        if self.ali_params:
            self.ali = Alignment(**self.ali_params)
        else:
            self.ali = Alignment(
                horizontal=self.ali_horizontal,
                vertical=self.ali_vertical,
                wrap_text=self.ali_wrap_text,
            )

        if self.border_params:
            self.border = Border(**self.border_params)
        else:
            self.border = Border(
                top=Side(
                    border_style=self.border_style_top,
                    color=self.border_color_top,
                ),
                right=Side(
                    border_style=self.border_style_right,
                    color=self.border_color_right,
                ),
                left=Side(
                    border_style=self.border_style_left,
                    color=self.border_color_left,
                ),
                bottom=Side(
                    border_style=self.border_style_bottom,
                    color=self.border_color_bottom,
                ),
            )

        if self.protect is True:
            self.protection = Protection(locked=True)
