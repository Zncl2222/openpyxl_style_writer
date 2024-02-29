import pytest
from openpyxl import load_workbook
from openpyxl_style_writer import RowWriter
from openpyxl.worksheet._write_only import WriteOnlyWorksheet


@pytest.mark.row_writer
class TestRowWriter:
    '''
    This test for openpyxl_style_writer is designed to ensure that the function can be executed
    without errors. However, it does not validate the correctness of the function's behavior.
    Additional test cases are needed to verify the actual functionality andbehavior of
    the function.
    '''

    sheet_name = 'test'

    def setup_method(self) -> None:
        self.rw = RowWriter()
        self.rw.create_sheet(self.sheet_name)
        self.rw.set_cell_width(1, 15)
        self.rw.set_cell_height(1, 25)

    def test_get_sheet(self) -> None:
        sheet = self.rw.get_current_sheet()
        assert isinstance(sheet, WriteOnlyWorksheet)

    def test_switch_sheet(self, capsys):
        sheet_name = 'test2'
        self.rw.create_sheet(sheet_name)
        self.rw.switch_current_sheet(sheet_name)
        print(self.rw.ws)
        captured = capsys.readouterr()
        expected_output = f'''<WriteOnlyWorksheet "{sheet_name}">\n'''
        assert captured.out == expected_output

    def test_row_append(self) -> None:
        self.rw.row_append('Hello')
        self.rw.create_row()

        self.rw.row_append(196632.124555, protect=True, number_format='#,##0.00')
        self.rw.create_row()

        saved_filename = 'test_output.xlsx'
        self.rw.save(saved_filename)

        wb = load_workbook(saved_filename)
        ws = wb[self.sheet_name]

        assert ws.cell(row=1, column=1).value == 'Hello'

    def test_row_append_list(self) -> None:
        data = ['Value 1', 'Value 2', 'Value 3']
        self.rw.row_append_list(data)
        self.rw.create_row()

        saved_filename = 'test_output.xlsx'
        self.rw.save(saved_filename)
        wb = load_workbook(saved_filename)
        ws = wb[self.sheet_name]

        for col, value in enumerate(data, start=1):
            assert ws.cell(row=1, column=col).value == value

    def test_print_row_list(self, capsys) -> None:
        data = ['Value 1', 'Value 2', 'Value 3']
        self.rw.row_append_list(data)
        self.rw.print_row_list()

        captured = capsys.readouterr()
        expected_output = 'row_list = [Value 1, Value 2, Value 3]\n'
        assert captured.out == expected_output

        self.rw.print_row_list(mode='')
        captured2 = capsys.readouterr()
        assert captured2.out == "[<Cell 'test'.A1>, <Cell 'test'.A1>, <Cell 'test'.A1>]\n"
